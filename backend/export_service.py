"""
Export Service for DataLogicEngine

Provides export functionality for:
- PDF export
- CSV export
- Excel export
"""

import io
import csv
import json
import logging
from datetime import datetime, UTC
from typing import Dict, Any, List

logger = logging.getLogger(__name__)


def export_simulation_to_csv(simulation: Any) -> io.StringIO:
    """
    Export simulation data to CSV format.
    
    Args:
        simulation: SimulationSession model instance
    
    Returns:
        StringIO object containing CSV data
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Field', 'Value'
    ])
    
    # Write simulation metadata
    writer.writerow(['ID', simulation.id])
    writer.writerow(['Session ID', simulation.session_id])
    writer.writerow(['Name', simulation.name])
    writer.writerow(['Description', simulation.description or ''])
    writer.writerow(['Status', simulation.status])
    writer.writerow(['Created At', simulation.created_at.isoformat() if simulation.created_at else ''])
    writer.writerow(['Started At', simulation.started_at.isoformat() if simulation.started_at else ''])
    writer.writerow(['Completed At', simulation.completed_at.isoformat() if simulation.completed_at else ''])
    writer.writerow(['Current Step', simulation.current_step])
    writer.writerow(['Total Steps', simulation.total_steps])
    
    # Write parameters
    writer.writerow([])
    writer.writerow(['Parameters', ''])
    if simulation.parameters:
        for key, value in simulation.parameters.items():
            writer.writerow([f'  {key}', str(value)])
    
    # Write results
    writer.writerow([])
    writer.writerow(['Results', ''])
    if simulation.results:
        for key, value in simulation.results.items():
            writer.writerow([f'  {key}', str(value)])
    
    output.seek(0)
    return output


def export_simulation_to_json(simulation: Any, pretty: bool = True) -> str:
    """
    Export simulation data to JSON format.
    
    Args:
        simulation: SimulationSession model instance
        pretty: Whether to format with indentation
    
    Returns:
        JSON string
    """
    data = {
        'id': simulation.id,
        'session_id': simulation.session_id,
        'name': simulation.name,
        'description': simulation.description,
        'status': simulation.status,
        'created_at': simulation.created_at.isoformat() if simulation.created_at else None,
        'started_at': simulation.started_at.isoformat() if simulation.started_at else None,
        'completed_at': simulation.completed_at.isoformat() if simulation.completed_at else None,
        'current_step': simulation.current_step,
        'total_steps': simulation.total_steps,
        'parameters': simulation.parameters,
        'results': simulation.results,
        'exported_at': datetime.now(UTC).isoformat()
    }
    
    if pretty:
        return json.dumps(data, indent=2)
    return json.dumps(data)


def export_nodes_to_csv(nodes: List[Any]) -> io.StringIO:
    """
    Export knowledge nodes to CSV format.
    
    Args:
        nodes: List of KnowledgeGraphNode instances
    
    Returns:
        StringIO object containing CSV data
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'ID', 'Node ID', 'Type', 'Label', 'Description', 
        'Axis Number', 'Created At', 'Updated At'
    ])
    
    # Write data
    for node in nodes:
        writer.writerow([
            node.id,
            node.node_id,
            node.node_type,
            node.label,
            node.description or '',
            node.axis_number,
            node.created_at.isoformat() if node.created_at else '',
            node.updated_at.isoformat() if node.updated_at else ''
        ])
    
    output.seek(0)
    return output


def generate_pdf_report(title: str, content: Dict[str, Any]) -> bytes:
    """
    Generate a PDF report.
    
    Note: Requires weasyprint for PDF generation.
    Falls back to HTML if weasyprint is not available.
    
    Args:
        title: Report title
        content: Report content dictionary
    
    Returns:
        PDF bytes or HTML string if weasyprint unavailable
    """
    # Generate HTML
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            h1 {{ color: #333; border-bottom: 2px solid #007bff; }}
            h2 {{ color: #555; margin-top: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #007bff; color: white; }}
            .meta {{ color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <p class="meta">Generated: {datetime.now(UTC).isoformat()}</p>
    """
    
    for section, data in content.items():
        html_content += f"<h2>{section}</h2>"
        if isinstance(data, dict):
            html_content += "<table>"
            for key, value in data.items():
                html_content += f"<tr><th>{key}</th><td>{value}</td></tr>"
            html_content += "</table>"
        elif isinstance(data, list):
            html_content += "<ul>"
            for item in data:
                html_content += f"<li>{item}</li>"
            html_content += "</ul>"
        else:
            html_content += f"<p>{data}</p>"
    
    html_content += """
    </body>
    </html>
    """
    
    try:
        from weasyprint import HTML
        # Generate PDF
        pdf = HTML(string=html_content).write_pdf()
        return pdf
        
    except ImportError:
        logger.warning("weasyprint not installed, PDF generation unavailable")
        # Return HTML as fallback
        return html_content.encode('utf-8')


def export_to_excel(data: List[Dict], sheet_name: str = "Data") -> io.BytesIO:
    """
    Export data to Excel format.
    
    Args:
        data: List of dictionaries to export
        sheet_name: Name of the Excel sheet
    
    Returns:
        BytesIO object containing Excel data
    """
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return io.BytesIO()
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except ImportError:
        logger.error("openpyxl not installed for Excel export")
        raise ImportError("openpyxl is required for Excel export")
